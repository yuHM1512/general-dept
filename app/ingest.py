from __future__ import annotations

from collections.abc import Callable
from typing import BinaryIO

from openpyxl import load_workbook
from sqlalchemy.dialects.postgresql import insert
from sqlmodel import Session

from app.models import PayrollRow
from app.services import classify_group, normalize_department, normalize_header, to_int_money


class IngestCounters:
    def __init__(self, *, total_rows: int, inserted: int, skipped: int, invalid: int):
        self.total_rows = total_rows
        self.inserted = inserted
        self.skipped = skipped
        self.invalid = invalid

    def merge(self, other: IngestCounters) -> None:
        """Accumulate counters from another sheet."""
        self.total_rows += other.total_rows
        self.inserted += other.inserted
        self.skipped += other.skipped
        self.invalid += other.invalid


# ── Format detection ─────────────────────────────────────────────────────────
# Legacy: single sheet with trailing space
TARGET_SHEET_NAME = "Luong ky nhan thang tong "

# New: one sheet per cơ sở, sheet name → co_so display value
MULTI_SHEET_MAP: dict[str, str] = {
    "Me Nhu": "Mẹ Nhu",
    "DT": "Duy Trung",
}


def _detect_sheets(wb) -> list[tuple[str, str | None]]:
    """Return list of (sheet_name, co_so_override) to ingest.

    - Legacy format: [("Luong ky nhan thang tong ", None)]
      → co_so derived from TTBP column at row level.
    - Multi-sheet format: [("Me Nhu", "Mẹ Nhu"), ("DT", "Duy Trung")]
      → co_so fixed per sheet.
    """
    if TARGET_SHEET_NAME in wb.sheetnames:
        return [(TARGET_SHEET_NAME, None)]

    found = [(sn, MULTI_SHEET_MAP[sn]) for sn in wb.sheetnames if sn in MULTI_SHEET_MAP]
    if found:
        return found

    raise ValueError(
        f"Không tìm thấy sheet hợp lệ. "
        f"Cần sheet '{TARGET_SHEET_NAME}' hoặc một trong {list(MULTI_SHEET_MAP.keys())}. "
        f"File có: {wb.sheetnames}"
    )


# ── Column resolution ────────────────────────────────────────────────────────

def _resolve_column_indices(ws) -> dict[str, int]:
    """Build 1-based column index map from header row.

    Returns keys: THANG, NAM, MANV, LGTRGIO, Bu du luong toi thieu,
    Tien CM thai 7T VSPN, Tien F L H R GL, TIEN E,
    TTBP, DEPARTMENT, FULL_NAME, and optionally JOB_TITLE, DON_VI.
    """
    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(1, col).value
        key = normalize_header(value)
        if key:
            headers[key] = col

    # ── Required money/identity columns (found by exact header name) ──
    required = {
        "THANG": "THANG",
        "NAM": "NAM",
        "MANV": "MANV",
        "LGTRGIO": "LGTRGIO",
        "Bu du luong toi thieu": "Bu du luong toi thieu",
        "Tien CM thai 7T VSPN": "Tien CM thai 7T VSPN",
        "Tien F L H R GL": "Tien F L H R GL",
        "TIEN E": "TIEN E",
    }

    col_index: dict[str, int] = {}
    missing = []
    for out_key, header_key in required.items():
        if header_key not in headers:
            missing.append(header_key)
            continue
        col_index[out_key] = headers[header_key]

    if missing:
        raise ValueError(f"Missing required headers: {', '.join(missing)}")

    # ── TTBP: try header "XN" (new format) or fallback col 4 (legacy) ──
    header_lc = {k.lower(): v for k, v in headers.items()}
    if "xn" in header_lc:
        col_index["TTBP"] = header_lc["xn"]
    else:
        col_index["TTBP"] = 4  # legacy positional

    # ── Department: always col 5 in both formats ──
    col_index["DEPARTMENT"] = 5

    # ── Full name: always col 7 in both formats ──
    col_index["FULL_NAME"] = 7

    # ── Job title: optional — only in legacy format (col 9) ──
    # Try common header names first; if none found, leave absent.
    for key in ["job_title", "chuc vu", "chức vụ", "chuc_vu"]:
        if key in header_lc:
            col_index["JOB_TITLE"] = header_lc[key]
            break

    # ── Don vi: optional explicit column ──
    for key in ["don_vi", "don vi", "đơn vị", "donvi"]:
        if key in header_lc:
            col_index["DON_VI"] = header_lc[key]
            break

    return col_index


# ── Batch size for bulk upsert ────────────────────────────────────────────────
_BATCH_SIZE = 2000


def ingest_workbook(file_obj: BinaryIO, session: Session) -> IngestCounters:
    return ingest_workbook_with_progress(file_obj, session, progress=None)


def ingest_workbook_with_progress(
    file_obj: BinaryIO,
    session: Session,
    *,
    progress: Callable[[int, int, int], None] | None,
) -> IngestCounters:
    wb = load_workbook(file_obj, data_only=True, read_only=True)
    sheets = _detect_sheets(wb)

    totals = IngestCounters(total_rows=0, inserted=0, skipped=0, invalid=0)
    for sheet_name, co_so_override in sheets:
        result = _ingest_sheet(
            wb[sheet_name],
            session,
            co_so_override=co_so_override,
            progress=progress,
            row_offset=totals.total_rows,
        )
        totals.merge(result)

    if progress:
        progress(totals.total_rows, totals.inserted, totals.invalid)
    return totals


def _ingest_sheet(
    ws,
    session: Session,
    *,
    co_so_override: str | None,
    progress: Callable[[int, int, int], None] | None,
    row_offset: int,
) -> IngestCounters:
    """Ingest a single worksheet into rcp_payrollrow."""
    # Resolve 1-based column indices, then convert to 0-based for tuple iteration.
    col1 = _resolve_column_indices(ws)
    c = {k: v - 1 for k, v in col1.items()}

    has_job_title = "JOB_TITLE" in c

    total_rows = 0
    inserted = 0
    invalid = 0
    batch: list[dict] = []

    def flush_batch() -> int:
        nonlocal batch
        if not batch:
            return 0
        stmt = insert(PayrollRow).values(batch)
        stmt = stmt.on_conflict_do_nothing(index_elements=["year", "month", "manv"])
        result = session.exec(stmt)
        session.commit()
        rowcount = int(getattr(result, "rowcount", 0) or 0)
        batch = []
        return rowcount

    row_iter = ws.iter_rows(min_row=2, values_only=True)
    for values in row_iter:
        total_rows += 1

        def _get(idx: int):
            return values[idx] if idx < len(values) else None

        manv = _get(c["MANV"])
        if manv is None or str(manv).strip() == "":
            invalid += 1
            if progress and ((row_offset + total_rows) % 500 == 0):
                progress(row_offset + total_rows, inserted, invalid)
            continue

        year = _get(c["NAM"])
        month = _get(c["THANG"])
        if year is None or month is None:
            invalid += 1
            if progress and ((row_offset + total_rows) % 500 == 0):
                progress(row_offset + total_rows, inserted, invalid)
            continue

        ttbp = _get(c["TTBP"]) or ""
        department = normalize_department(_get(c["DEPARTMENT"]))
        full_name = _get(c["FULL_NAME"]) or ""
        job_title = _get(c["JOB_TITLE"]) if has_job_title else ""
        job_title = job_title or ""

        # co_so: nếu có override từ sheet name thì dùng, không thì suy từ TTBP
        if co_so_override is not None:
            co_so = co_so_override
        else:
            co_so = "Duy Trung" if str(ttbp).strip() == "DT" else "Mẹ Nhu"

        don_vi_idx = c.get("DON_VI")
        don_vi = _get(don_vi_idx) if don_vi_idx is not None else None
        if don_vi is None:
            don_vi = str(ttbp).strip()

        lgtrgio = to_int_money(_get(c["LGTRGIO"]))
        bu = to_int_money(_get(c["Bu du luong toi thieu"]))
        cm = to_int_money(_get(c["Tien CM thai 7T VSPN"]))
        flhr = to_int_money(_get(c["Tien F L H R GL"]))
        tien_e = to_int_money(_get(c["TIEN E"]))

        metric = lgtrgio + bu + cm + flhr + tien_e
        group_name = classify_group(str(department), str(job_title))

        manv_str = str(manv).strip()
        try:
            year_int = int(float(year))
            month_int = int(float(month))
        except (TypeError, ValueError):
            invalid += 1
            continue

        batch.append(
            {
                "year": year_int,
                "month": month_int,
                "ttbp": str(ttbp).strip(),
                "don_vi": str(don_vi).strip(),
                "co_so": co_so,
                "department": department,
                "manv": manv_str,
                "full_name": str(full_name).strip(),
                "job_title": str(job_title).strip(),
                "lgtrgio": lgtrgio,
                "bu_du_luong_toi_thieu": bu,
                "tien_cm_thai_7t_vspn": cm,
                "tien_f_l_h_r_gl": flhr,
                "tien_e": tien_e,
                "metric_vnd": metric,
                "group_name": group_name,
            }
        )
        if len(batch) >= _BATCH_SIZE:
            inserted += flush_batch()
            if progress:
                progress(row_offset + total_rows, inserted, invalid)
        elif progress and ((row_offset + total_rows) % 500 == 0):
            progress(row_offset + total_rows, inserted, invalid)

    inserted += flush_batch()

    valid = total_rows - invalid
    skipped = max(0, valid - inserted)
    return IngestCounters(total_rows=total_rows, inserted=inserted, skipped=skipped, invalid=invalid)
