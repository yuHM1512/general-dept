from __future__ import annotations

import io
import threading
from pathlib import Path
from datetime import date, datetime
from uuid import UUID
from concurrent.futures import ThreadPoolExecutor
from urllib.parse import quote

from fastapi import Depends, FastAPI, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy import func
from sqlalchemy import exists
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlmodel import Session, select
from starlette.middleware.sessions import SessionMiddleware
from starlette.requests import Request

from app.db import create_db_and_tables, get_session
from app.ingest import ingest_workbook_with_progress
from app.models import GeneralEmployee, HangingLine, IngestJob, PayrollRow
from app.schemas import (
    IngestJobResponse,
    IngestJobStatus,
    FilterOptionsResponse,
    MonthlyDetailsResponse,
    ParticipationResponse,
    HeadcountResponse,
    InsightsResponse,
    PayrollListResponse,
    PayrollRowOut,
    StatsResponse,
    TimeseriesResponse,
    EmployeeBelowTargetResponse,
    EmployeeBelowTargetItem,
    BelowTargetCountResponse,
    BelowTargetBreakdownResponse,
    BelowTargetDonutItem,
    BelowTargetTableRow,
    HeadcountUniqueResponse,
    HangingLineItem,
    HangingLineListResponse,
)
from app.settings import settings
from app.stats import month_stats, year_timeseries
from app.stats import headcount_by_month, metric_insights

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

app = FastAPI(title=settings.app_name)
executor = ThreadPoolExecutor(max_workers=2)
CO_SO_OPTIONS = ["Mẹ Nhu", "Duy Trung"]

BASE_DIR = Path(__file__).resolve().parent.parent
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))
static_dir = BASE_DIR / "static"
static_dir.mkdir(exist_ok=True)
app.mount("/static", StaticFiles(directory=str(static_dir)), name="static")

uploads_dir = BASE_DIR / "data" / "uploads"
uploads_dir.mkdir(parents=True, exist_ok=True)


def _current_user(request: Request) -> dict | None:
    try:
        return request.session.get("user")  # type: ignore[attr-defined]
    except Exception:
        return None


@app.middleware("http")
async def _require_login_for_rcp(request: Request, call_next):
    path = request.url.path

    if path == "/" or path == "/health" or path.startswith("/static"):
        return await call_next(request)

    if path.startswith("/rcp") or path.startswith("/api"):
        if path.startswith("/rcp/login"):
            return await call_next(request)

        user = _current_user(request)
        if not user:
            if path.startswith("/api"):
                return JSONResponse({"detail": "Not authenticated"}, status_code=401)

            next_url = path
            if request.url.query:
                next_url = f"{path}?{request.url.query}"
            return RedirectResponse(url=f"/rcp/login?next={quote(next_url)}", status_code=303)

    return await call_next(request)


app.add_middleware(
    SessionMiddleware,
    secret_key=settings.session_secret,
    same_site="lax",
    https_only=False,
)


def _ensure_no_active_ingest(session: Session) -> None:
    active = session.exec(
        select(func.count())
        .select_from(IngestJob)
        .where(IngestJob.status.in_(["pending", "running"]))
    ).one()
    if int(active or 0) > 0:
        raise HTTPException(status_code=409, detail="Ingest is still running. Please wait until it completes.")


def _apply_hanging_only(query, *, hanging_only: bool) -> any:
    if not hanging_only:
        return query
    sub = (
        select(HangingLine.id)
        .where(HangingLine.don_vi == PayrollRow.don_vi)
        .where(HangingLine.department == PayrollRow.department)
        .limit(1)
    )
    return query.where(exists(sub))


def _split_multi(value: str | None) -> list[str] | None:
    if not value:
        return None
    parts = [p.strip() for p in value.split(",") if p.strip()]
    return parts or None


def _apply_csv_in(query, column, csv_value: str | None):
    values = _split_multi(csv_value)
    if not values:
        return query
    # Some source values contain stray spaces; use trim() for string columns so
    # filters match what the UI displays (labels are already stripped).
    try:
        py_type = getattr(getattr(column, "type", None), "python_type", None)
    except Exception:
        py_type = None
    col_expr = func.trim(column) if py_type is str else column
    return query.where(col_expr.in_(values))


@app.on_event("startup")
def _startup() -> None:
    if settings.create_tables_on_startup:
        create_db_and_tables()


@app.get("/health")
def health() -> dict:
    return {"status": "ok", "env": settings.app_env}


@app.get("/", response_class=HTMLResponse)
def homepage(request: Request) -> HTMLResponse:
    # Department homepage (portal)
    return templates.TemplateResponse(
        "homepage.html",
        {
            "request": request,
            "app_name": settings.app_name,
            "target_salary_vnd_fmt": _fmt_vnd(settings.target_salary_vnd),
            "now_year": datetime.utcnow().year,
            "active_nav": "homepage",
            "user": _current_user(request),
            "department_name": "Phòng Tổng hợp",
        },
    )


@app.get("/rcp/login", response_class=HTMLResponse)
def rcp_login_page(
    request: Request,
    next: str | None = Query(default=None),
    session: Session = Depends(get_session),
) -> HTMLResponse:
    # Ensure tables exist so login can work even on a fresh DB.
    create_db_and_tables()

    if _current_user(request):
        return RedirectResponse(url=next or "/rcp", status_code=303)

    return templates.TemplateResponse(
        "login_rcp.html",
        {
            "request": request,
            "app_name": settings.app_name,
            "now_year": datetime.utcnow().year,
            "target_salary_vnd_fmt": _fmt_vnd(settings.target_salary_vnd),
            "next_url": next or "/rcp",
            "error": None,
        },
    )


@app.post("/rcp/login")
def rcp_login_submit(
    request: Request,
    ma_nv: str = Form(...),
    next: str | None = Query(default=None),
    session: Session = Depends(get_session),
):
    create_db_and_tables()

    code = (ma_nv or "").strip().upper()
    if not code:
        return templates.TemplateResponse(
            "login_rcp.html",
            {
                "request": request,
                "app_name": settings.app_name,
                "now_year": datetime.utcnow().year,
                "target_salary_vnd_fmt": _fmt_vnd(settings.target_salary_vnd),
                "next_url": next or "/rcp",
                "error": "Vui lòng nhập mã nhân viên.",
            },
            status_code=400,
        )

    employee = session.get(GeneralEmployee, code)
    if not employee:
        return templates.TemplateResponse(
            "login_rcp.html",
            {
                "request": request,
                "app_name": settings.app_name,
                "now_year": datetime.utcnow().year,
                "target_salary_vnd_fmt": _fmt_vnd(settings.target_salary_vnd),
                "next_url": next or "/rcp",
                "error": "Mã nhân viên không hợp lệ hoặc chưa được cấp quyền.",
            },
            status_code=401,
        )

    request.session["user"] = {  # type: ignore[attr-defined]
        "ma_nv": employee.ma_nv,
        "ho_ten": employee.ho_ten,
        "chuc_vu": employee.chuc_vu,
        "don_vi": employee.don_vi,
        "bo_phan": employee.bo_phan,
    }

    return RedirectResponse(url=next or "/rcp", status_code=303)


@app.get("/rcp/logout")
def rcp_logout(request: Request) -> RedirectResponse:
    try:
        request.session.clear()  # type: ignore[attr-defined]
    except Exception:
        pass
    return RedirectResponse(url="/rcp/login", status_code=303)


@app.get("/rcp", response_class=HTMLResponse)
def rcp_home(request: Request) -> HTMLResponse:
    return templates.TemplateResponse(
        "home_rcp.html",
        {
            "request": request,
            "app_name": settings.app_name,
            "target_salary_vnd_fmt": _fmt_vnd(settings.target_salary_vnd),
            "now_year": datetime.utcnow().year,
            "active_nav": "rcp",
            "user": _current_user(request),
        },
    )


@app.get("/rcp/data", response_class=HTMLResponse)
def data_page(request: Request) -> HTMLResponse:
    return templates.TemplateResponse(
        "data_rcp.html",
        {
            "request": request,
            "app_name": settings.app_name,
            "target_salary_vnd_fmt": _fmt_vnd(settings.target_salary_vnd),
            "now_year": datetime.utcnow().year,
            "active_nav": "rcp_data",
            "user": _current_user(request),
        },
    )


@app.get("/rcp/dashboard", response_class=HTMLResponse)
def dashboard_page(request: Request) -> HTMLResponse:
    return templates.TemplateResponse(
        "dashboard_rcp.html",
        {
            "request": request,
            "app_name": settings.app_name,
            "target_salary_vnd_fmt": _fmt_vnd(settings.target_salary_vnd),
            "now_year": datetime.utcnow().year,
            "active_nav": "rcp_dashboard",
            "target_salary_vnd": settings.target_salary_vnd,
            "user": _current_user(request),
        },
    )


@app.get("/rcp/preview/below-target", response_class=HTMLResponse)
def preview_below_target_page(
    request: Request,
    year: int = Query(..., ge=2000, le=2100),
    month: int | None = Query(None, ge=1, le=12),
    view: str = Query("month", pattern="^(month|year_avg|year_sum)$"),
    group_name: str | None = Query(None),
    co_so: str | None = Query(None),
    don_vi: str | None = Query(None),
    department: str | None = Query(None),
    exclude_don_vi: str | None = Query(None, description="CSV, exact match"),
    exclude_department: str | None = Query(None, description="CSV, exact match"),
    hanging_only: bool = Query(False),
    session: Session = Depends(get_session),
) -> HTMLResponse:
    """
    A print-friendly, continuous preview (PDF-like) of employees below target.
    Intended to be opened from the dashboard modal via "Xem đầy đủ".
    """
    _ensure_no_active_ingest(session)
    if view in ("year_avg", "year_sum"):
        month = None
    elif month is None:
        raise HTTPException(status_code=400, detail="month is required when view=month")

    max_rows = int(settings.preview_max_rows or 20000)
    dept_lc = func.lower(PayrollRow.department)
    ex_dv = _split_multi(exclude_don_vi)
    ex_dep = _split_multi(exclude_department)

    def _common_filters(q):
        q = q.where(~(dept_lc.like("%lãnh đạo%") | dept_lc.like("%lanh dao%")))
        q = _apply_hanging_only(q, hanging_only=hanging_only)
        q = _apply_csv_in(q, PayrollRow.group_name, group_name)
        if co_so:
            q = q.where(PayrollRow.co_so == co_so)
        q = _apply_csv_in(q, PayrollRow.don_vi, don_vi)
        q = _apply_csv_in(q, PayrollRow.department, department)
        if ex_dv:
            q = q.where(~func.trim(PayrollRow.don_vi).in_(ex_dv))
        if ex_dep:
            q = q.where(~func.trim(PayrollRow.department).in_(ex_dep))
        return q

    rows: list[EmployeeBelowTargetItem] = []
    total = 0
    truncated = False

    if month is not None:
        base = select(func.count()).select_from(PayrollRow).where(
            PayrollRow.year == year,
            PayrollRow.month == month,
            PayrollRow.metric_vnd < settings.target_salary_vnd,
        )
        base = _common_filters(base)
        total = int(session.exec(base).one() or 0)

        q = (
            select(
                PayrollRow.manv,
                PayrollRow.full_name,
                PayrollRow.don_vi,
                PayrollRow.department,
                PayrollRow.metric_vnd,
            )
            .where(
                PayrollRow.year == year,
                PayrollRow.month == month,
                PayrollRow.metric_vnd < settings.target_salary_vnd,
            )
        )
        q = _common_filters(q)
        q = q.order_by(PayrollRow.metric_vnd.asc(), PayrollRow.full_name.asc()).limit(max_rows + 1)
        raw = session.exec(q).all()
        if len(raw) > max_rows:
            truncated = True
            raw = raw[:max_rows]
        rows = [
            EmployeeBelowTargetItem(
                manv=str(r.manv or ""),
                full_name=str(r.full_name or ""),
                don_vi=str(r.don_vi or ""),
                department=str(r.department or ""),
                month=int(month or 0),
                metric_vnd=int(r.metric_vnd or 0),
            )
            for r in raw
        ]
    elif view == "year_avg":
        grouped = (
            select(
                PayrollRow.manv.label("manv"),
                func.max(PayrollRow.full_name).label("full_name"),
                func.max(PayrollRow.don_vi).label("don_vi"),
                func.max(PayrollRow.department).label("department"),
                func.avg(PayrollRow.metric_vnd).label("avg_vnd"),
            )
            .where(PayrollRow.year == year)
        )
        grouped = _common_filters(grouped)
        grouped = grouped.group_by(PayrollRow.manv).having(func.avg(PayrollRow.metric_vnd) < settings.target_salary_vnd)
        subq = grouped.subquery()
        total = int(session.exec(select(func.count()).select_from(subq)).one() or 0)

        page = session.exec(
            select(subq.c.manv, subq.c.full_name, subq.c.don_vi, subq.c.department, subq.c.avg_vnd)
            .order_by(subq.c.avg_vnd.asc(), subq.c.full_name.asc())
            .limit(max_rows + 1)
        ).all()
        if len(page) > max_rows:
            truncated = True
            page = page[:max_rows]
        rows = [
            EmployeeBelowTargetItem(
                manv=str(r.manv or ""),
                full_name=str(r.full_name or ""),
                don_vi=str(r.don_vi or ""),
                department=str(r.department or ""),
                month=None,
                metric_vnd=int(round(float(r.avg_vnd or 0.0))),
            )
            for r in page
        ]
    else:
        q = (
            select(
                PayrollRow.manv,
                PayrollRow.full_name,
                PayrollRow.don_vi,
                PayrollRow.department,
                PayrollRow.month,
                PayrollRow.metric_vnd,
            )
            .where(PayrollRow.year == year, PayrollRow.metric_vnd < settings.target_salary_vnd)
        )
        q = _common_filters(q)
        total = int(session.exec(select(func.count()).select_from(q.subquery())).one() or 0)
        raw = session.exec(
            q.order_by(PayrollRow.month.asc(), PayrollRow.metric_vnd.asc(), PayrollRow.full_name.asc()).limit(max_rows + 1)
        ).all()
        if len(raw) > max_rows:
            truncated = True
            raw = raw[:max_rows]
        rows = [
            EmployeeBelowTargetItem(
                manv=str(r.manv or ""),
                full_name=str(r.full_name or ""),
                don_vi=str(r.don_vi or ""),
                department=str(r.department or ""),
                month=int(r.month or 0) if r.month is not None else None,
                metric_vnd=int(r.metric_vnd or 0),
            )
            for r in raw
        ]

    def _scope_hint() -> str:
        if month is not None:
            return f"tháng {month}/{year}"
        if view == "year_sum":
            return f"năm {year} (tổng các tháng)"
        return f"năm {year} (bình quân theo nhân sự)"

    return templates.TemplateResponse(
        "preview_below_target.html",
        {
            "request": request,
            "app_name": settings.app_name,
            "now_year": datetime.utcnow().year,
            "year": year,
            "month": month,
            "target_salary_vnd": settings.target_salary_vnd,
            "target_salary_vnd_fmt": _fmt_vnd(settings.target_salary_vnd),
            "scope_hint": _scope_hint(),
            "filters": {
                "view": view,
                "group_name": group_name,
                "co_so": co_so,
                "don_vi": don_vi,
                "department": department,
                "hanging_only": hanging_only,
                "exclude_don_vi": exclude_don_vi,
                "exclude_department": exclude_department,
            },
            "total": total,
            "shown": len(rows),
            "truncated": truncated,
            "max_rows": max_rows,
            "rows": rows,
            "user": _current_user(request),
        },
    )


@app.get("/data")
def legacy_data_redirect() -> RedirectResponse:
    return RedirectResponse(url="/rcp/data", status_code=307)


@app.get("/dashboard")
def legacy_dashboard_redirect() -> RedirectResponse:
    return RedirectResponse(url="/rcp/dashboard", status_code=307)


def _fmt_vnd(value: int) -> str:
    try:
        return f"{int(value):,}".replace(",", ".")
    except Exception:
        return str(value)


def _run_ingest_job(job_id: UUID, xlsx_path: Path) -> None:
    from sqlmodel import Session

    from app.db import engine

    with Session(engine) as session:
        job = session.get(IngestJob, job_id)
        if not job:
            return
        job.status = "running"
        job.started_at = datetime.utcnow()
        session.add(job)
        session.commit()

        try:
            # Set total rows early for UI.
            try:
                from openpyxl import load_workbook

                wb = load_workbook(xlsx_path, data_only=True, read_only=True)
                if "Luong ky nhan thang tong " in wb.sheetnames:
                    expected_total = max(0, wb["Luong ky nhan thang tong "].max_row - 1)
                    job.total_rows = expected_total
                    session.add(job)
                    session.commit()
            except Exception:
                # Non-fatal: progress can still work without total_rows.
                pass

            def _progress(processed_rows: int, inserted: int, invalid: int) -> None:
                # Persist progress using a separate session to avoid interfering with bulk inserts.
                try:
                    with Session(engine) as s2:
                        current = s2.get(IngestJob, job_id)
                        if not current:
                            return
                        current.processed_rows = processed_rows
                        current.inserted = inserted
                        current.invalid = invalid
                        s2.add(current)
                        s2.commit()
                except Exception:
                    return

            with xlsx_path.open("rb") as f:
                counters = ingest_workbook_with_progress(f, session, progress=_progress)
            job.total_rows = counters.total_rows
            job.processed_rows = counters.total_rows
            job.inserted = counters.inserted
            job.skipped = counters.skipped
            job.invalid = counters.invalid
            job.status = "completed"
            job.finished_at = datetime.utcnow()
            job.error = None
            session.add(job)
            session.commit()
        except Exception as exc:
            job.status = "failed"
            job.finished_at = datetime.utcnow()
            job.error = str(exc)
            session.add(job)
            session.commit()


@app.post("/api/ingest/excel", response_model=IngestJobResponse)
async def ingest_excel(
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
) -> IngestJobResponse:
    # Ensure tables exist even if CREATE_TABLES_ON_STARTUP is off.
    create_db_and_tables()
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Please upload a .xlsx file")

    content = await file.read()
    job = IngestJob(status="pending")
    session.add(job)
    session.commit()
    session.refresh(job)

    xlsx_path = uploads_dir / f"{job.id}.xlsx"
    xlsx_path.write_bytes(content)

    executor.submit(_run_ingest_job, job.id, xlsx_path)
    return IngestJobResponse(job_id=str(job.id))


@app.post("/api/ingest/local", response_model=IngestJobResponse)
def ingest_local(
    path: str = Query(..., description="Local path to .xlsx"),
    session: Session = Depends(get_session),
) -> IngestJobResponse:
    if not settings.allow_local_ingest:
        raise HTTPException(status_code=403, detail="Local ingest disabled (ALLOW_LOCAL_INGEST=false)")
    create_db_and_tables()
    xlsx = Path(path)
    if not xlsx.exists() or xlsx.suffix.lower() != ".xlsx":
        raise HTTPException(status_code=400, detail="Invalid path")
    job = IngestJob(status="pending")
    session.add(job)
    session.commit()
    session.refresh(job)
    executor.submit(_run_ingest_job, job.id, xlsx)
    return IngestJobResponse(job_id=str(job.id))


@app.get("/api/ingest/jobs/{job_id}", response_model=IngestJobStatus)
def ingest_job_status(job_id: str, session: Session = Depends(get_session)) -> IngestJobStatus:
    create_db_and_tables()
    try:
        uid = UUID(job_id)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Invalid job_id") from exc
    job = session.get(IngestJob, uid)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    return IngestJobStatus(
        job_id=str(job.id),
        status=job.status,
        total_rows=job.total_rows,
        processed_rows=job.processed_rows,
        inserted=job.inserted,
        skipped=job.skipped,
        invalid=job.invalid,
        error=job.error,
    )


@app.get("/api/stats", response_model=StatsResponse)
def stats(
    year: int = Query(..., ge=2000, le=2100),
    month: int = Query(..., ge=1, le=12),
    session: Session = Depends(get_session),
) -> StatsResponse:
    _ensure_no_active_ingest(session)
    stats_rows = month_stats(session, year=year, month=month, target_salary_vnd=settings.target_salary_vnd)
    return StatsResponse(year=year, month=month, target_salary_vnd=settings.target_salary_vnd, stats=stats_rows)


@app.get("/api/timeseries", response_model=TimeseriesResponse)
def timeseries(
    year: int = Query(..., ge=2000, le=2100),
    session: Session = Depends(get_session),
) -> TimeseriesResponse:
    _ensure_no_active_ingest(session)
    points = year_timeseries(session, year=year, target_salary_vnd=settings.target_salary_vnd)
    return TimeseriesResponse(year=year, target_salary_vnd=settings.target_salary_vnd, points=points)


@app.get("/api/insights", response_model=InsightsResponse)
def insights(
    year: int = Query(..., ge=2000, le=2100),
    month: int | None = Query(None, ge=1, le=12),
    view: str = Query("month", pattern="^(month|year_avg)$"),
    group_name: str | None = Query(None),
    co_so: str | None = Query(None),
    don_vi: str | None = Query(None),
    department: str | None = Query(None),
    hanging_only: bool = Query(False),
    session: Session = Depends(get_session),
) -> InsightsResponse:
    _ensure_no_active_ingest(session)
    if view == "year_avg":
        month = None

    by_group, by_month = metric_insights(
        session,
        year=year,
        month=month,
        target_salary_vnd=settings.target_salary_vnd,
        group_filter=group_name,
        co_so=co_so,
        don_vi=don_vi,
        department=department,
        hanging_only=hanging_only,
    )

    return InsightsResponse(
        year=year,
        month=month,
        target_salary_vnd=settings.target_salary_vnd,
        view=view,
        group_filter=group_name,
        by_group=by_group,
        by_month=by_month,
    )


@app.get("/api/headcount", response_model=HeadcountResponse)
def headcount(
    year: int = Query(..., ge=2000, le=2100),
    group_name: str | None = Query(None),
    co_so: str | None = Query(None),
    don_vi: str | None = Query(None),
    department: str | None = Query(None),
    hanging_only: bool = Query(False),
    session: Session = Depends(get_session),
) -> HeadcountResponse:
    _ensure_no_active_ingest(session)
    points = headcount_by_month(
        session,
        year=year,
        group_filter=group_name,
        co_so=co_so,
        don_vi=don_vi,
        department=department,
        hanging_only=hanging_only,
    )
    avg = int(round(sum(p["headcount"] for p in points) / len(points))) if points else 0
    return HeadcountResponse(year=year, group_filter=group_name, points=points, avg_headcount=avg)




@app.get("/api/available-months")
def available_months(session: Session = Depends(get_session)) -> list[dict]:
    rows = session.exec(select(PayrollRow.year, PayrollRow.month).distinct().order_by(PayrollRow.year, PayrollRow.month)).all()
    return [{"year": y, "month": m} for (y, m) in rows]


@app.get("/api/trend/monthly-details", response_model=MonthlyDetailsResponse)
def monthly_trend_details(
    year: int = Query(..., ge=2000, le=2100),
    month: int = Query(..., ge=1, le=12),
    group_name: str | None = Query(None),
    co_so: str | None = Query(None),
    don_vi: str | None = Query(None),
    department: str | None = Query(None),
    hanging_only: bool = Query(False),
    session: Session = Depends(get_session),
) -> MonthlyDetailsResponse:
    """
    Drilldown for a month: Đơn vị -> Bộ phận, with avg metric_vnd.
    Uses the same exclusion rule for lãnh đạo departments only.
    """
    _ensure_no_active_ingest(session)

    query = (
        select(
            PayrollRow.don_vi.label("don_vi"),
            PayrollRow.department.label("department"),
            func.count().label("count"),
            func.avg(PayrollRow.metric_vnd).label("avg_vnd"),
        )
        .where(PayrollRow.year == year, PayrollRow.month == month)
    )

    dept_lc = func.lower(PayrollRow.department)
    query = query.where(
        ~(
            dept_lc.like("%lãnh đạo%")
            | dept_lc.like("%lanh dao%")
        )
    )
    query = _apply_csv_in(query, PayrollRow.group_name, group_name)
    if co_so:
        query = query.where(PayrollRow.co_so == co_so)
    query = _apply_csv_in(query, PayrollRow.don_vi, don_vi)
    query = _apply_csv_in(query, PayrollRow.department, department)
    query = _apply_hanging_only(query, hanging_only=hanging_only)

    query = query.group_by(PayrollRow.don_vi, PayrollRow.department).order_by(PayrollRow.don_vi, PayrollRow.department)
    rows = session.exec(query).all()

    # Build hierarchical structure: don_vi -> departments
    by_dv: dict[str, dict] = {}
    for r in rows:
        dv = (r.don_vi or "").strip()
        dep = (r.department or "").strip()
        count = int(r.count or 0)
        avg_vnd = float(r.avg_vnd or 0.0)

        if dv not in by_dv:
            by_dv[dv] = {"sum": 0.0, "count": 0, "departments": []}
        by_dv[dv]["sum"] += avg_vnd * count
        by_dv[dv]["count"] += count
        by_dv[dv]["departments"].append({"department": dep, "avg_vnd": int(round(avg_vnd))})

    items = []
    for dv in sorted(by_dv.keys()):
        total_count = int(by_dv[dv]["count"] or 0)
        weighted_avg = int(round((by_dv[dv]["sum"] / total_count) if total_count else 0.0))
        items.append(
            {
                "don_vi": dv,
                "avg_vnd": weighted_avg,
                "departments": by_dv[dv]["departments"],
            }
        )

    return MonthlyDetailsResponse(year=year, month=month, items=items)


@app.get("/api/participation", response_model=ParticipationResponse)
def participation(
    year: int = Query(..., ge=2000, le=2100),
    month: int | None = Query(None, ge=1, le=12),
    group_name: str | None = Query(None),
    co_so: str | None = Query(None),
    don_vi: str | None = Query(None),
    department: str | None = Query(None),
    hanging_only: bool = Query(False),
    session: Session = Depends(get_session),
) -> ParticipationResponse:
    """
    Count how many distinct `don_vi` participate in the selected scope.
    Excludes departments that contain "lãnh đạo" (case-insensitive).
    """
    _ensure_no_active_ingest(session)
    query = select(func.count(func.distinct(PayrollRow.don_vi))).where(PayrollRow.year == year)
    if month is not None:
        query = query.where(PayrollRow.month == month)

    dept_lc = func.lower(PayrollRow.department)
    query = query.where(~(dept_lc.like("%lãnh đạo%") | dept_lc.like("%lanh dao%")))

    query = _apply_csv_in(query, PayrollRow.group_name, group_name)
    if co_so:
        query = query.where(PayrollRow.co_so == co_so)
    query = _apply_csv_in(query, PayrollRow.don_vi, don_vi)
    query = _apply_csv_in(query, PayrollRow.department, department)
    query = _apply_hanging_only(query, hanging_only=hanging_only)

    count = session.exec(query).one()
    return ParticipationResponse(year=year, month=month, don_vi_count=int(count or 0))


@app.get("/api/below-target-count", response_model=BelowTargetCountResponse)
def below_target_count(
    year: int = Query(..., ge=2000, le=2100),
    month: int | None = Query(None, ge=1, le=12),
    view: str = Query("month", pattern="^(month|year_avg)$"),
    group_name: str | None = Query(None),
    co_so: str | None = Query(None),
    don_vi: str | None = Query(None),
    department: str | None = Query(None),
    hanging_only: bool = Query(False),
    session: Session = Depends(get_session),
) -> BelowTargetCountResponse:
    _ensure_no_active_ingest(session)
    if view == "year_avg":
        month = None

    dept_lc = func.lower(PayrollRow.department)

    if month is not None:
        q = select(func.count(func.distinct(PayrollRow.manv))).where(
            PayrollRow.year == year,
            PayrollRow.month == month,
            PayrollRow.metric_vnd < settings.target_salary_vnd,
        )
        q = q.where(~(dept_lc.like("%lãnh đạo%") | dept_lc.like("%lanh dao%")))
        q = _apply_csv_in(q, PayrollRow.group_name, group_name)
        if co_so:
            q = q.where(PayrollRow.co_so == co_so)
        q = _apply_csv_in(q, PayrollRow.don_vi, don_vi)
        q = _apply_csv_in(q, PayrollRow.department, department)
        q = _apply_hanging_only(q, hanging_only=hanging_only)
        below = int(session.exec(q).one() or 0)
        return BelowTargetCountResponse(
            year=year,
            month=month,
            view="month",
            target_salary_vnd=settings.target_salary_vnd,
            below_count=below,
        )

    # year_avg (dashboard month=all): sum of monthly distinct employees below target (not average).
    monthly = (
        select(
            PayrollRow.month.label("month"),
            func.count(func.distinct(PayrollRow.manv)).label("cnt"),
        )
        .where(PayrollRow.year == year, PayrollRow.metric_vnd < settings.target_salary_vnd)
    )
    monthly = monthly.where(~(dept_lc.like("%lãnh đạo%") | dept_lc.like("%lanh dao%")))
    monthly = _apply_csv_in(monthly, PayrollRow.group_name, group_name)
    if co_so:
        monthly = monthly.where(PayrollRow.co_so == co_so)
    monthly = _apply_csv_in(monthly, PayrollRow.don_vi, don_vi)
    monthly = _apply_csv_in(monthly, PayrollRow.department, department)
    monthly = _apply_hanging_only(monthly, hanging_only=hanging_only)
    monthly = monthly.group_by(PayrollRow.month)
    rows = session.exec(monthly).all()
    below_total = int(sum(int(r.cnt or 0) for r in rows))  # type: ignore[attr-defined]
    return BelowTargetCountResponse(
        year=year,
        month=None,
        view="year_avg",
        target_salary_vnd=settings.target_salary_vnd,
        below_count=below_total,
    )


@app.get("/api/employees/below-target", response_model=EmployeeBelowTargetResponse)
def employees_below_target(
    year: int = Query(..., ge=2000, le=2100),
    month: int | None = Query(None, ge=1, le=12),
    view: str = Query("month", pattern="^(month|year_avg|year_sum)$"),
    group_name: str | None = Query(None),
    co_so: str | None = Query(None),
    don_vi: str | None = Query(None),
    department: str | None = Query(None),
    exclude_don_vi: str | None = Query(None, description="CSV, exact match"),
    exclude_department: str | None = Query(None, description="CSV, exact match"),
    hanging_only: bool = Query(False),
    limit: int = Query(50, ge=1, le=500),
    offset: int = Query(0, ge=0),
    session: Session = Depends(get_session),
) -> EmployeeBelowTargetResponse:
    _ensure_no_active_ingest(session)
    if view in ("year_avg", "year_sum"):
        month = None
    elif month is None:
        raise HTTPException(status_code=400, detail="month is required when view=month")

    dept_lc = func.lower(PayrollRow.department)
    ex_dv = _split_multi(exclude_don_vi)
    ex_dep = _split_multi(exclude_department)

    if month is not None:
        base = select(PayrollRow).where(
            PayrollRow.year == year,
            PayrollRow.month == month,
            PayrollRow.metric_vnd < settings.target_salary_vnd,
        )
        base = base.where(~(dept_lc.like("%lãnh đạo%") | dept_lc.like("%lanh dao%")))
        base = _apply_hanging_only(base, hanging_only=hanging_only)
        base = _apply_csv_in(base, PayrollRow.group_name, group_name)
        if co_so:
            base = base.where(PayrollRow.co_so == co_so)
        base = _apply_csv_in(base, PayrollRow.don_vi, don_vi)
        base = _apply_csv_in(base, PayrollRow.department, department)
        if ex_dv:
            base = base.where(~func.trim(PayrollRow.don_vi).in_(ex_dv))
        if ex_dep:
            base = base.where(~func.trim(PayrollRow.department).in_(ex_dep))

        total = int(session.exec(select(func.count()).select_from(base.subquery())).one() or 0)
        q = (
            select(
                PayrollRow.manv,
                PayrollRow.full_name,
                PayrollRow.don_vi,
                PayrollRow.department,
                PayrollRow.metric_vnd,
            )
            .where(
                PayrollRow.year == year,
                PayrollRow.month == month,
                PayrollRow.metric_vnd < settings.target_salary_vnd,
            )
        )
        q = q.where(~(dept_lc.like("%lãnh đạo%") | dept_lc.like("%lanh dao%")))
        q = _apply_hanging_only(q, hanging_only=hanging_only)
        q = _apply_csv_in(q, PayrollRow.group_name, group_name)
        if co_so:
            q = q.where(PayrollRow.co_so == co_so)
        q = _apply_csv_in(q, PayrollRow.don_vi, don_vi)
        q = _apply_csv_in(q, PayrollRow.department, department)
        if ex_dv:
            q = q.where(~func.trim(PayrollRow.don_vi).in_(ex_dv))
        if ex_dep:
            q = q.where(~func.trim(PayrollRow.department).in_(ex_dep))
        q = q.order_by(PayrollRow.metric_vnd.asc(), PayrollRow.full_name.asc()).offset(offset).limit(limit)
        rows = session.exec(q).all()

        items = [
            EmployeeBelowTargetItem(
                manv=str(r.manv or ""),
                full_name=str(r.full_name or ""),
                don_vi=str(r.don_vi or ""),
                department=str(r.department or ""),
                month=int(month or 0),
                metric_vnd=int(r.metric_vnd or 0),
            )
            for r in rows
        ]
        return EmployeeBelowTargetResponse(
            year=year,
            month=month,
            view="month",
            target_salary_vnd=settings.target_salary_vnd,
            total=total,
            limit=limit,
            offset=offset,
            rows=items,
        )

    if view == "year_avg":
        grouped = (
            select(
                PayrollRow.manv.label("manv"),
                func.max(PayrollRow.full_name).label("full_name"),
                func.max(PayrollRow.don_vi).label("don_vi"),
                func.max(PayrollRow.department).label("department"),
                func.avg(PayrollRow.metric_vnd).label("avg_vnd"),
            )
            .where(PayrollRow.year == year)
        )
        grouped = grouped.where(~(dept_lc.like("%lãnh đạo%") | dept_lc.like("%lanh dao%")))
        grouped = _apply_hanging_only(grouped, hanging_only=hanging_only)
        grouped = _apply_csv_in(grouped, PayrollRow.group_name, group_name)
        if co_so:
            grouped = grouped.where(PayrollRow.co_so == co_so)
        grouped = _apply_csv_in(grouped, PayrollRow.don_vi, don_vi)
        grouped = _apply_csv_in(grouped, PayrollRow.department, department)
        if ex_dv:
            grouped = grouped.where(~func.trim(PayrollRow.don_vi).in_(ex_dv))
        if ex_dep:
            grouped = grouped.where(~func.trim(PayrollRow.department).in_(ex_dep))
        grouped = grouped.group_by(PayrollRow.manv).having(func.avg(PayrollRow.metric_vnd) < settings.target_salary_vnd)
        subq = grouped.subquery()
        total = int(session.exec(select(func.count()).select_from(subq)).one() or 0)
        page = session.exec(
            select(subq.c.manv, subq.c.full_name, subq.c.don_vi, subq.c.department, subq.c.avg_vnd)
            .order_by(subq.c.avg_vnd.asc(), subq.c.full_name.asc())
            .offset(offset)
            .limit(limit)
        ).all()
        items = [
            EmployeeBelowTargetItem(
                manv=str(r.manv or ""),
                full_name=str(r.full_name or ""),
                don_vi=str(r.don_vi or ""),
                department=str(r.department or ""),
                month=None,
                metric_vnd=int(round(float(r.avg_vnd or 0.0))),
            )
            for r in page
        ]
        return EmployeeBelowTargetResponse(
            year=year,
            month=None,
            view="year_avg",
            target_salary_vnd=settings.target_salary_vnd,
            total=total,
            limit=limit,
            offset=offset,
            rows=items,
        )

    # view == "year_sum": list employee-month rows below target (so totals match donut/table sums).
    q = (
        select(
            PayrollRow.manv,
            PayrollRow.full_name,
            PayrollRow.don_vi,
            PayrollRow.department,
            PayrollRow.month,
            PayrollRow.metric_vnd,
        )
        .where(PayrollRow.year == year, PayrollRow.metric_vnd < settings.target_salary_vnd)
    )
    q = q.where(~(dept_lc.like("%lãnh đạo%") | dept_lc.like("%lanh dao%")))
    q = _apply_hanging_only(q, hanging_only=hanging_only)
    q = _apply_csv_in(q, PayrollRow.group_name, group_name)
    if co_so:
        q = q.where(PayrollRow.co_so == co_so)
    q = _apply_csv_in(q, PayrollRow.don_vi, don_vi)
    q = _apply_csv_in(q, PayrollRow.department, department)
    if ex_dv:
        q = q.where(~func.trim(PayrollRow.don_vi).in_(ex_dv))
    if ex_dep:
        q = q.where(~func.trim(PayrollRow.department).in_(ex_dep))

    total = int(session.exec(select(func.count()).select_from(q.subquery())).one() or 0)
    rows = session.exec(q.order_by(PayrollRow.month.asc(), PayrollRow.metric_vnd.asc(), PayrollRow.full_name.asc()).offset(offset).limit(limit)).all()
    items = [
        EmployeeBelowTargetItem(
            manv=str(r.manv or ""),
            full_name=str(r.full_name or ""),
            don_vi=str(r.don_vi or ""),
            department=str(r.department or ""),
            month=int(r.month or 0) if r.month is not None else None,
            metric_vnd=int(r.metric_vnd or 0),
        )
        for r in rows
    ]
    return EmployeeBelowTargetResponse(
        year=year,
        month=None,
        view="year_sum",
        target_salary_vnd=settings.target_salary_vnd,
        total=total,
        limit=limit,
        offset=offset,
        rows=items,
    )


@app.get("/api/headcount-unique", response_model=HeadcountUniqueResponse)
def headcount_unique(
    year: int = Query(..., ge=2000, le=2100),
    group_name: str | None = Query(None),
    co_so: str | None = Query(None),
    don_vi: str | None = Query(None),
    department: str | None = Query(None),
    hanging_only: bool = Query(False),
    session: Session = Depends(get_session),
) -> HeadcountUniqueResponse:
    _ensure_no_active_ingest(session)
    dept_lc = func.lower(PayrollRow.department)
    q = select(func.count(func.distinct(PayrollRow.manv))).where(PayrollRow.year == year)
    q = q.where(~(dept_lc.like("%lãnh đạo%") | dept_lc.like("%lanh dao%")))
    q = _apply_csv_in(q, PayrollRow.group_name, group_name)
    if co_so:
        q = q.where(PayrollRow.co_so == co_so)
    q = _apply_csv_in(q, PayrollRow.don_vi, don_vi)
    q = _apply_csv_in(q, PayrollRow.department, department)
    q = _apply_hanging_only(q, hanging_only=hanging_only)
    headcount = int(session.exec(q).one() or 0)
    return HeadcountUniqueResponse(year=year, headcount=headcount)


@app.get("/api/below-target/breakdown", response_model=BelowTargetBreakdownResponse)
def below_target_breakdown(
    year: int = Query(..., ge=2000, le=2100),
    month: int | None = Query(None, ge=1, le=12),
    view: str = Query("month", pattern="^(month|year_avg)$"),
    group_name: str | None = Query(None),
    co_so: str | None = Query(None),
    don_vi: str | None = Query(None),
    department: str | None = Query(None),
    hanging_only: bool = Query(False),
    session: Session = Depends(get_session),
) -> BelowTargetBreakdownResponse:
    """
    Breakdown of employees below target:
    - If don_vi is not selected: donut shows counts by don_vi
    - If don_vi is selected: donut shows top 5 departments within that don_vi

    Count semantics:
    - view=month: distinct employees below target in the selected month
    - view=year_avg (month=None): average monthly distinct employees below target
    Sparkline always shows month-by-month counts within the selected year.
    """
    _ensure_no_active_ingest(session)
    if view == "year_avg":
        month = None
    elif month is None:
        raise HTTPException(status_code=400, detail="month is required when view=month")

    dept_lc = func.lower(PayrollRow.department)

    # Build a month-by-month matrix for the whole year (to support sparklines + month selection).
    q = (
        select(
            PayrollRow.month.label("month"),
            func.trim(PayrollRow.don_vi).label("don_vi"),
            func.trim(PayrollRow.department).label("department"),
            func.count(func.distinct(PayrollRow.manv)).label("cnt"),
        )
        .where(PayrollRow.year == year, PayrollRow.metric_vnd < settings.target_salary_vnd)
    )
    q = q.where(~(dept_lc.like("%lãnh đạo%") | dept_lc.like("%lanh dao%")))
    q = _apply_hanging_only(q, hanging_only=hanging_only)
    q = _apply_csv_in(q, PayrollRow.group_name, group_name)
    if co_so:
        q = q.where(PayrollRow.co_so == co_so)
    q = _apply_csv_in(q, PayrollRow.don_vi, don_vi)
    q = _apply_csv_in(q, PayrollRow.department, department)
    q = q.group_by(PayrollRow.month, PayrollRow.don_vi, PayrollRow.department)

    rows = session.exec(q).all()

    months = sorted({int(r.month or 0) for r in rows if int(r.month or 0) > 0})  # type: ignore[attr-defined]
    if not months:
        return BelowTargetBreakdownResponse(
            year=year,
            month=month,
            view=view if month is not None else "year_avg",
            target_salary_vnd=settings.target_salary_vnd,
            mode="department" if don_vi else "don_vi",
            months=[],
            total=0,
            donut=[],
            table=[],
        )

    month_index = {m: i for i, m in enumerate(months)}

    # Table matrix by (don_vi, department)
    matrix: dict[tuple[str, str], list[int]] = {}
    for r in rows:
        m = int(r.month or 0)  # type: ignore[attr-defined]
        if m not in month_index:
            continue
        dv = (r.don_vi or "").strip()  # type: ignore[attr-defined]
        dep = (r.department or "").strip()  # type: ignore[attr-defined]
        key = (dv, dep)
        if key not in matrix:
            matrix[key] = [0 for _ in months]
        matrix[key][month_index[m]] = int(r.cnt or 0)  # type: ignore[attr-defined]

    # Helper to derive a single "count" value for month view.
    def _month_value(vals: list[int]) -> int:
        idx = month_index.get(month or 0)
        return int(vals[idx]) if idx is not None else 0

    # Donut
    # - If exactly 1 đơn vị is selected: show top departments within that đơn vị
    # - Otherwise: show distribution by đơn vị (even if filtered to multiple đơn vị)
    don_vi_values = _split_multi(don_vi)
    donut_mode = "department" if (don_vi_values and len(don_vi_values) == 1) else "don_vi"
    donut_counts: dict[str, list[int]] = {}
    for (dv, dep), vals in matrix.items():
        key = dep if donut_mode == "department" else dv
        if key not in donut_counts:
            donut_counts[key] = [0 for _ in months]
        donut_counts[key] = [a + b for a, b in zip(donut_counts[key], vals)]

    if month is not None:
        donut_items_all = [(label, _month_value(vals)) for label, vals in donut_counts.items()]
        donut_items_all = [(l, c) for (l, c) in donut_items_all if c > 0]
        donut_items_all.sort(key=lambda x: x[1], reverse=True)
        total_all = int(sum(c for (_l, c) in donut_items_all))
        alloc_map = {str(l): int(c) for (l, c) in donut_items_all}
    else:
        # month=all: sum across all months (not average) to align with count.
        donut_items_all = [(label, int(sum(vals))) for label, vals in donut_counts.items()]
        donut_items_all = [(l, c) for (l, c) in donut_items_all if c > 0]
        donut_items_all.sort(key=lambda x: x[1], reverse=True)
        total_all = int(sum(c for (_l, c) in donut_items_all))
        alloc_map = {str(l): int(c) for (l, c) in donut_items_all}

    limit = 5 if donut_mode == "department" else 8
    donut_top = donut_items_all[:limit]
    sum_top = int(sum(c for (_l, c) in donut_top))
    remainder = max(0, total_all - sum_top)
    donut_items = donut_top + ([("Khác", remainder)] if remainder > 0 else [])

    donut = [BelowTargetDonutItem(label=(l or "(Trống)"), count=int(c)) for (l, c) in donut_items if int(c) > 0]
    total = total_all

    # Table rows (top 25)
    table_rows: list[BelowTargetTableRow] = []
    if month is None:
        for (dv, dep), vals in matrix.items():
            c = int(sum(vals))
            if c <= 0:
                continue
            table_rows.append(
                BelowTargetTableRow(
                    don_vi=dv or "(Trống)",
                    department=dep or "(Trống)",
                    count=c,
                    sparkline=[int(v) for v in vals],
                )
            )
    else:
        for (dv, dep), vals in matrix.items():
            c = _month_value(vals)
            if c <= 0:
                continue
            table_rows.append(
                BelowTargetTableRow(
                    don_vi=dv or "(Trống)",
                    department=dep or "(Trống)",
                    count=int(c),
                    sparkline=[int(v) for v in vals],
                )
            )
    # IMPORTANT:
    # The UI groups table rows by `don_vi`. If we truncate by department rows first,
    # totals per `don_vi` will be undercounted and won't match the donut counts.
    # So we keep full department rows, then take top 25 `don_vi` by total, and
    # return all departments within those top `don_vi`.
    dv_totals: dict[str, int] = {}
    for r in table_rows:
        dv_totals[r.don_vi] = dv_totals.get(r.don_vi, 0) + int(r.count or 0)
    top_dv = {dv for (dv, _c) in sorted(dv_totals.items(), key=lambda x: x[1], reverse=True)[:25]}
    table_rows = [r for r in table_rows if r.don_vi in top_dv]
    table_rows.sort(key=lambda r: r.count, reverse=True)

    return BelowTargetBreakdownResponse(
        year=year,
        month=month,
        view="month" if month is not None else "year_avg",
        target_salary_vnd=settings.target_salary_vnd,
        mode=donut_mode,
        months=months,
        total=total,
        donut=donut,
        table=table_rows,
    )


@app.get("/api/filters", response_model=FilterOptionsResponse)
def filter_options(
    year: int = Query(..., ge=2000, le=2100),
    group_name: str | None = Query(None),
    co_so: str | None = Query(None),
    don_vi: str | None = Query(None),
    department: str | None = Query(None),
    hanging_only: bool = Query(False),
    session: Session = Depends(get_session),
) -> FilterOptionsResponse:
    _ensure_no_active_ingest(session)

    # Faceted filter options:
    # - Months: apply all selected filters (except month itself; month not a param here)
    # - Each dimension's option list ignores its own current selection so users can switch
    #   without having to reset to "Tất cả" first (e.g., co_so list still shows both).
    def _base() -> any:
        q = select(PayrollRow).where(PayrollRow.year == year)
        dept_lc = func.lower(PayrollRow.department)
        q = q.where(
            ~(
                dept_lc.like("%lãnh đạo%")
                | dept_lc.like("%lanh dao%")
            )
        )
        q = _apply_hanging_only(q, hanging_only=hanging_only)
        q = _apply_csv_in(q, PayrollRow.group_name, group_name)
        return q

    def _distinct_list(q: any, field_name: str) -> list:
        subq = q.subquery()
        col = getattr(subq.c, field_name)
        return session.exec(select(col).distinct().order_by(col)).all()

    # months uses all selections
    q_months = _base()
    if co_so:
        q_months = q_months.where(PayrollRow.co_so == co_so)
    q_months = _apply_csv_in(q_months, PayrollRow.don_vi, don_vi)
    q_months = _apply_csv_in(q_months, PayrollRow.department, department)
    months = _distinct_list(q_months, "month")

    # co_so is a derived dimension with only two values.
    # Always return both so users can switch directly without resetting other filters.
    co_so_values = CO_SO_OPTIONS

    # don_vi ignores don_vi selection
    q_don_vi = _base()
    if co_so:
        q_don_vi = q_don_vi.where(PayrollRow.co_so == co_so)
    q_don_vi = _apply_csv_in(q_don_vi, PayrollRow.department, department)
    don_vi_values = _distinct_list(q_don_vi, "don_vi")

    # departments ignores department selection
    q_dept = _base()
    if co_so:
        q_dept = q_dept.where(PayrollRow.co_so == co_so)
    q_dept = _apply_csv_in(q_dept, PayrollRow.don_vi, don_vi)
    dept_values = _distinct_list(q_dept, "department")

    return FilterOptionsResponse(
        year=year,
        months=[int(m) for m in months if m is not None],
        co_so=[str(x) for x in co_so_values if x],
        don_vi=[str(x) for x in don_vi_values if x],
        departments=[str(x) for x in dept_values if x],
    )


@app.get("/api/payroll", response_model=PayrollListResponse)
def payroll_rows(
    year: int = Query(..., ge=2000, le=2100),
    month: int = Query(..., ge=1, le=12),
    limit: int = Query(50, ge=1, le=200),
    offset: int = Query(0, ge=0),
    group_name: str | None = Query(None),
    manv: str | None = Query(None, description="Contains"),
    department: str | None = Query(None, description="Contains"),
    co_so: str | None = Query(None),
    don_vi: str | None = Query(None),
    hanging_only: bool = Query(False),
    session: Session = Depends(get_session),
) -> PayrollListResponse:
    _ensure_no_active_ingest(session)
    query = select(PayrollRow).where(PayrollRow.year == year, PayrollRow.month == month)
    dept_lc = func.lower(PayrollRow.department)
    query = query.where(
        ~(
            dept_lc.like("%lãnh đạo%")
            | dept_lc.like("%lanh dao%")
        )
    )
    query = _apply_csv_in(query, PayrollRow.group_name, group_name)
    if co_so:
        query = query.where(PayrollRow.co_so == co_so)
    query = _apply_csv_in(query, PayrollRow.don_vi, don_vi)
    query = _apply_hanging_only(query, hanging_only=hanging_only)
    if manv:
        query = query.where(PayrollRow.manv.ilike(f"%{manv.strip()}%"))
    # department supports multi-select via CSV (exact match). If not multi, keep "contains" behavior.
    dept_values = _split_multi(department)
    if dept_values and len(dept_values) > 1:
        query = query.where(PayrollRow.department.in_(dept_values))
    elif department:
        query = query.where(PayrollRow.department.ilike(f"%{department.strip()}%"))

    total = session.exec(select(func.count()).select_from(query.subquery())).one()
    rows = session.exec(query.order_by(PayrollRow.manv).offset(offset).limit(limit)).all()
    out = [
        PayrollRowOut(
            year=r.year,
            month=r.month,
            ttbp=r.ttbp,
            don_vi=r.don_vi,
            co_so=r.co_so,
            department=r.department,
            manv=r.manv,
            full_name=r.full_name,
            job_title=r.job_title,
            group_name=r.group_name,
            metric_vnd=r.metric_vnd,
        )
        for r in rows
    ]
    return PayrollListResponse(year=year, month=month, total=int(total or 0), limit=limit, offset=offset, rows=out)


@app.get("/api/debug/db")
def debug_db(session: Session = Depends(get_session)) -> dict:
    """
    Lightweight diagnostics for connectivity + whether any rows were ingested.
    """
    create_db_and_tables()
    row_count = session.exec(select(func.count()).select_from(PayrollRow)).one()
    # Avoid returning rows; just counts and distinct months.
    months = session.exec(
        select(PayrollRow.year, PayrollRow.month).distinct().order_by(PayrollRow.year, PayrollRow.month)
    ).all()
    return {"ok": True, "row_count": int(row_count or 0), "months": [{"year": y, "month": m} for (y, m) in months]}


@app.get("/api/hanging-lines", response_model=HangingLineListResponse)
def list_hanging_lines(
    don_vi: str | None = Query(None),
    session: Session = Depends(get_session),
) -> HangingLineListResponse:
    """
    Danh sách bộ phận đã được khai báo là "chuyền treo".
    """
    create_db_and_tables()
    q = select(HangingLine)
    if don_vi:
        q = q.where(HangingLine.don_vi == don_vi)
    q = q.order_by(HangingLine.don_vi, HangingLine.department, HangingLine.id)
    items = session.exec(q).all()
    return HangingLineListResponse(
        items=[
            HangingLineItem(
                id=int(x.id or 0),
                don_vi=x.don_vi,
                department=x.department,
                ngay_ap_dung=x.ngay_ap_dung.isoformat(),
                created_at=x.created_at.isoformat(),
            )
            for x in items
        ]
    )


@app.post("/api/hanging-lines", response_model=HangingLineItem)
def add_hanging_line(
    don_vi: str = Form(...),
    department: str = Form(...),
    ngay_ap_dung: str | None = Form(None),
    session: Session = Depends(get_session),
) -> HangingLineItem:
    """
    Thêm 1 bộ phận vào danh sách "chuyền treo" theo cặp (don_vi, department).
    Chống trùng theo unique index.
    """
    create_db_and_tables()
    dv = (don_vi or "").strip()
    dep = (department or "").strip()
    if not dv or not dep:
        raise HTTPException(status_code=422, detail="don_vi và department là bắt buộc")

    now = datetime.utcnow()
    applied = date.today()
    if ngay_ap_dung and ngay_ap_dung.strip():
        try:
            applied = date.fromisoformat(ngay_ap_dung.strip())
        except Exception:
            raise HTTPException(status_code=422, detail="ngay_ap_dung không hợp lệ (YYYY-MM-DD)")
    t = HangingLine.__table__
    stmt = (
        pg_insert(t)
        .values(don_vi=dv, department=dep, ngay_ap_dung=applied, created_at=now)
        .on_conflict_do_nothing(index_elements=["don_vi", "department"])
        .returning(t.c.id, t.c.don_vi, t.c.department, t.c.ngay_ap_dung, t.c.created_at)
    )
    row = session.exec(stmt).first()
    if row is None:
        # Already exists; fetch existing row
        existing = session.exec(
            select(HangingLine).where(HangingLine.don_vi == dv, HangingLine.department == dep).limit(1)
        ).first()
        if existing is None:
            raise HTTPException(status_code=500, detail="Không thể tạo hanging_line")
        return HangingLineItem(
            id=int(existing.id or 0),
            don_vi=existing.don_vi,
            department=existing.department,
            ngay_ap_dung=existing.ngay_ap_dung.isoformat(),
            created_at=existing.created_at.isoformat(),
        )

    # row is a sqlalchemy Row, access by attribute/index
    hid, hdv, hdep, happlied, hcreated = row
    session.commit()
    return HangingLineItem(
        id=int(hid),
        don_vi=str(hdv),
        department=str(hdep),
        ngay_ap_dung=happlied.isoformat() if happlied else applied.isoformat(),
        created_at=hcreated.isoformat() if hcreated else datetime.utcnow().isoformat(),
    )


@app.delete("/api/hanging-lines/{line_id}")
def delete_hanging_line(
    line_id: int,
    session: Session = Depends(get_session),
) -> dict:
    create_db_and_tables()
    item = session.get(HangingLine, line_id)
    if not item:
        raise HTTPException(status_code=404, detail="Not found")
    session.delete(item)
    session.commit()
    return {"ok": True, "deleted_id": line_id}


@app.get("/api/export/below-target.xlsx")
def export_below_target_excel(
    year: int = Query(..., ge=2000, le=2100),
    month: int | None = Query(None, ge=1, le=12),
    group_name: str | None = Query(None),
    co_so: str | None = Query(None),
    don_vi: str | None = Query(None),
    department: str | None = Query(None),
    hanging_only: bool = Query(False),
    session: Session = Depends(get_session),
):
    """
    Export employees below target to Excel.
    Semantics:
    - If month is provided: rows are employee-month for that month
    - If month is None: rows are employee-month across the whole year (sum-of-months view)
    """
    _ensure_no_active_ingest(session)

    dept_lc = func.lower(PayrollRow.department)
    q = (
        select(
            PayrollRow.full_name,
            PayrollRow.manv,
            PayrollRow.don_vi,
            PayrollRow.department,
            PayrollRow.month,
            PayrollRow.metric_vnd,
        )
        .where(PayrollRow.year == year, PayrollRow.metric_vnd < settings.target_salary_vnd)
        .where(~(dept_lc.like("%lãnh đạo%") | dept_lc.like("%lanh dao%")))
    )
    if month is not None:
        q = q.where(PayrollRow.month == month)
    q = _apply_hanging_only(q, hanging_only=hanging_only)
    q = _apply_csv_in(q, PayrollRow.group_name, group_name)
    if co_so:
        q = q.where(PayrollRow.co_so == co_so)
    q = _apply_csv_in(q, PayrollRow.don_vi, don_vi)
    q = _apply_csv_in(q, PayrollRow.department, department)
    q = q.order_by(PayrollRow.month.asc(), PayrollRow.don_vi.asc(), PayrollRow.department.asc(), PayrollRow.metric_vnd.asc(), PayrollRow.full_name.asc())

    rows = session.exec(q).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Duoi muc tieu"

    header = ["Họ tên", "Mã nhân viên", "Đơn vị", "Bộ phận", "Tháng", "Mức lương"]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0B3B91")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in rows:
        mm = int(r.month or 0)
        ym = f"{year}-{mm:02d}" if mm else str(year)
        ws.append(
            [
                str(r.full_name or ""),
                str(r.manv or ""),
                str((r.don_vi or "").strip()),
                str((r.department or "").strip()),
                ym,
                int(r.metric_vnd or 0),
            ]
        )

    # Formatting
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 14
    for cell in ws["F"][1:]:
        cell.number_format = "#,##0"
        cell.alignment = Alignment(horizontal="right")
    for row_idx in range(2, ws.max_row + 1):
        ws[f"E{row_idx}"].alignment = Alignment(horizontal="center")

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    m_label = f"{month:02d}" if month is not None else "all"
    filename = f"duoi_muc_tieu_{year}_{m_label}.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
